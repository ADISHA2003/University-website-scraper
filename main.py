import requests
from bs4 import BeautifulSoup
import openpyxl
import re

# Function to fetch the HTML content from a URL
def get_html(url):
    response = requests.get(url)
    response.raise_for_status()  # Raise an error for bad responses
    return response.text

# Function to scrape university data
def scrape_university_data(soup):
    university_name = scrape_university_name(soup)
    university_logo = scrape_university_logo(soup)
    university_picture = scrape_university_picture(soup)
    university_url = scrape_university_url(soup)
    contact_number, full_address = scrape_contact_info(soup)

    return {
        'University Name': university_name,
        'University Logo': university_logo,
        'University Picture': university_picture,
        'University URL': university_url,
        'Contact Number': contact_number,
        'Full Address': full_address
    }

# Function to scrape university name
def scrape_university_name(soup):
    h1_tag = soup.find('h1', class_='page-title h1')
    return h1_tag.text.strip() if h1_tag else "University name not found"

# Function to scrape university logo
def scrape_university_logo(soup):
    logo_tag = soup.find('link', rel='icon')
    return logo_tag['href'] if logo_tag and 'href' in logo_tag.attrs else "University logo not found"

# Function to scrape university picture
def scrape_university_picture(soup):
    image_div = soup.find('div', class_='c-image__image')
    if image_div and 'style' in image_div.attrs:
        match = re.search(r'url\(["\']?([^"\')]+)["\']?\)', image_div['style'])
        return match.group(1) if match else "University picture not found"
    return "University picture not found"

# Function to scrape university URL
def scrape_university_url(soup):
    meta_tag = soup.find('meta', property='og:url')
    return meta_tag['content'] if meta_tag and 'content' in meta_tag.attrs else "University URL not found"

# Function to scrape university contact info
def scrape_contact_info(soup):
    contact_info_div = soup.find('div', class_='s-sink t-sink c-subheader__desc')
    if contact_info_div:
        paragraphs = contact_info_div.find_all('p')
        contact_number = ""
        full_address = ""
        for p in paragraphs:
            text = p.get_text(strip=True)
            if text.startswith("For general inquiries"):
                contact_number = text.split("call ")[1]
            elif "Mailing address:" in text:
                full_address = "\n".join([p.get_text(strip=True) for p in paragraphs[paragraphs.index(p)+1:]])
        return contact_number, full_address
    return "Contact number not found", "Address not found"

# Function to scrape application process
def scrape_application_process(soup):
    application_process = []
    sections = ['how-to-apply-for-aid', 'prospective-students', 'current-students', 'financial-aid-forms']
    for section_id in sections:
        section = soup.find(id=section_id)
        if section:
            section_text = section.get_text(strip=True)
            application_process.append({'section': section_id.replace('-', ' ').title(), 'content': section_text})
    return application_process

# Function to scrape course data
def scrape_course_data(soup):
    courses_data = []
    course_teasers = soup.find_all('article', class_='content-type--course')
    for course in course_teasers:
        course_title = course.find('h3').text.strip()
        course_description = course.find('div', class_='field--name-field-teaser-description').text.strip()
        course_tagline = course.find('div', class_='field--name-field-tagline').text.strip()
        course_link = course.find('a', class_='read-more-link')['href']
        course_info = {
            'Title': course_title,
            'Tagline': course_tagline,
            'Description': course_description,
            'Link': course_link
        }
        courses_data.append(course_info)
    return courses_data

# Function to save data to Excel
def save_to_excel(university_data, application_process, courses_data, file_name="university_data.xlsx"):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "University Info"

    # University data
    headers = ['University Name', 'University Logo', 'University Picture', 'University URL', 'Contact Number', 'Full Address']
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header)
    for col, value in enumerate(university_data.values(), start=1):
        sheet.cell(row=2, column=col, value=value)

    # Application process data
    sheet.append([])  # Empty row before application process
    sheet.append(['Application Process'])
    for item in application_process:
        sheet.append([item['section'], item['content']])

    # Course data
    sheet.append([])  # Empty row before course data
    sheet.append(['Courses'])
    course_headers = ['Title', 'Tagline', 'Description', 'Link']
    sheet.append(course_headers)
    for course in courses_data:
        sheet.append([course['Title'], course['Tagline'], course['Description'], course['Link']])

    # Save the workbook
    workbook.save(file_name)
    print(f"Data saved to {file_name}")

# Main function to perform scraping and saving
def main():
    university_url = 'https://www.harvard.edu/'  # Replace with actual university URL
    financial_aid_url = 'https://college.harvard.edu/financial-aid/apply-financial-aid'
    courses_url = "https://www.harvardonline.harvard.edu/"

    # Scrape university data
    html_content = get_html(university_url)
    soup = BeautifulSoup(html_content, 'html.parser')
    university_data = scrape_university_data(soup)

    # Scrape application process
    html_content = get_html(financial_aid_url)
    soup = BeautifulSoup(html_content, 'html.parser')
    application_process = scrape_application_process(soup)

    # Scrape course data
    html_content = get_html(courses_url)
    soup = BeautifulSoup(html_content, 'html.parser')
    courses_data = scrape_course_data(soup)

    # Save all the data to an Excel file
    save_to_excel(university_data, application_process, courses_data)

# Run the script
main()
